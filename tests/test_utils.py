from unittest.mock import patch

from openpyxl import Workbook

from src.utils import read_excel


def create_mock_workbook():
    # Создаем новый рабочий лист в памяти
    wb = Workbook()
    ws = wb.active
    # Добавляем заголовки и данные
    ws.append(['Header1', 'Header2', 'Header3'])
    ws.append(['Value1', 'Value2', 'Value3'])
    ws.append(['Value4', 'Value5', 'Value6'])

    return wb


@patch('openpyxl.load_workbook')
def test_read_excel(mock_load_workbook):
    # Настраиваем mock для возврата нашего фиктивного рабочего листа
    mock_wb = create_mock_workbook()
    mock_load_workbook.return_value = mock_wb
    # Ожидаемый результат
    expected_data = [
        {'Header1': 'Value1', 'Header2': 'Value2', 'Header3': 'Value3'},
        {'Header1': 'Value4', 'Header2': 'Value5', 'Header3': 'Value6'}
    ]

    # Вызываем тестируемую функцию
    result = read_excel()

    # Проверяем, что результат соответствует ожиданиям
    assert result == expected_data
